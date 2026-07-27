# -*- coding: utf-8 -*-
"""
Полная выгрузка данных ГТЛ из текущего проекта QGIS в Excel — версия 4.

Скрипт:
- только читает загруженные в QGIS слои и таблицы;
- НЕ изменяет PostgreSQL;
- формирует новый XLSX с 11 рабочими листами и листом контроля;
- использует подтверждённые UUID-связи;
- объединяет последовательные SegmentWorkPeriods в маршруты.

Подтверждённая логика:
- название месторождения или полевого объекта берётся из GtsPoints.Name;
- GUID объекта берётся из FieldPoints.Id;
- номер лицензии берётся только из LicUch.lic;
- пустые LuId, SubsoilUserId и Collectors.FieldPointId допустимы;
- периоды без дат идут в лист "Режим работы";
- периоды с датами идут в лист "ЗимаЛето";
- маршруты строятся по всем направлениям графа; одинаковые начала и концы допустимы;
- при WorkType=1 направление сегментов разворачивается;
- для нулевых сегментов в служебной колонке перемычек ставится 1.

Запуск:
1. Откройте проект QGIS, в котором загружены таблицы из PostgreSQL.
2. Откройте: Модули -> Консоль Python -> Показать редактор.
3. Откройте этот файл и нажмите "Запустить сценарий".
4. Выберите место сохранения XLSX.
"""

import math
import os
import re
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path

from qgis.core import QgsProject, QgsMessageLog, Qgis
from qgis.PyQt.QtCore import QDate, QDateTime, QUuid, QStandardPaths
from qgis.PyQt.QtWidgets import QFileDialog, QMessageBox

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError as exc:
    QMessageBox.critical(
        None,
        "Не найден openpyxl",
        "В Python QGIS не установлена библиотека openpyxl.\n\n"
        "Скрипт не запускался и база не изменялась.\n"
        "Перед повторным запуском необходимо установить openpyxl "
        "в окружение Python, используемое QGIS."
    )
    raise


LOG_TAG = "GTL Excel Export"

# Названия таблиц/слоёв, ожидаемые в текущем проекте QGIS.
EXPECTED_LAYERS = [
    "BalancePoints",
    "Collectors",
    "CompressionStations",
    "ConnectionPoints",
    "FieldPointAssignments",
    "FieldPoints",
    "GasDeliveryPoints",
    "GasDistributionStations",
    "GasMeasuringStations",
    "GasPipeSegments",
    "GasPipelines",
    "GtsPoints",
    "InZones",
    "LicUch",
    "OutZones",
    "Regions",
    "SegmentWorkPeriods",
    "SubsoilUsers",
    "TransGases",
    "UndergroundGasStorages",
]

# Критичные слои, без которых основную структуру собрать нельзя.
CRITICAL_LAYERS = [
    "GtsPoints",
    "GasPipeSegments",
    "GasPipelines",
    "Regions",
]

# Кэш фактических имён полей с учётом регистра.
_FIELD_MAP_CACHE = {}


def log(message, level=Qgis.Info):
    QgsMessageLog.logMessage(str(message), LOG_TAG, level)


def clean_value(value):
    """Преобразует QVariant/QDate/QUuid и NULL к обычным Python-значениям."""
    if value is None:
        return None

    if isinstance(value, QDateTime):
        return value.toPyDateTime()

    if isinstance(value, QDate):
        return value.toPyDate()

    if isinstance(value, QUuid):
        return value.toString().strip("{}")

    if isinstance(value, float) and math.isnan(value):
        return None

    text = str(value).strip()
    if text.upper() in {"NULL", "NONE", "<NULL>"}:
        return None

    return value


def uuid_key(value):
    """Нормализованный ключ UUID/идентификатора для словарей."""
    value = clean_value(value)
    if value is None:
        return None
    return str(value).strip().strip("{}").lower()


def get_layer(name):
    """Ищет слой сначала точно, затем без учёта регистра."""
    project = QgsProject.instance()

    exact = project.mapLayersByName(name)
    if exact:
        return exact[0]

    wanted = name.casefold()
    for layer in project.mapLayers().values():
        if layer.name().casefold() == wanted:
            return layer

    return None


def get_field_map(layer):
    if layer is None:
        return {}

    layer_id = layer.id()
    if layer_id not in _FIELD_MAP_CACHE:
        _FIELD_MAP_CACHE[layer_id] = {
            field.name().casefold(): field.name()
            for field in layer.fields()
        }
    return _FIELD_MAP_CACHE[layer_id]


def attr(layer, feature, *candidates, default=None):
    """Получает атрибут по одному из возможных имён, без учёта регистра."""
    if layer is None or feature is None:
        return default

    field_map = get_field_map(layer)
    for candidate in candidates:
        actual = field_map.get(candidate.casefold())
        if actual is not None:
            value = clean_value(feature[actual])
            return default if value is None else value

    return default


def build_index(layer, *id_candidates):
    """Индексирует слой по UUID/ID."""
    result = {}
    if layer is None:
        return result

    candidates = id_candidates or ("Id", "id")
    for feature in layer.getFeatures():
        feature_id = attr(layer, feature, *candidates)
        key = uuid_key(feature_id)
        if key:
            result[key] = feature
    return result


def lookup_feature(index, value):
    return index.get(uuid_key(value))


def as_float(value):
    value = clean_value(value)
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).replace(" ", "").replace(",", "."))
    except (TypeError, ValueError):
        return None


def is_zero(value, tolerance=1e-12):
    number = as_float(value)
    return number is not None and abs(number) <= tolerance


def bool_ru(value):
    value = clean_value(value)
    if value is None:
        return None
    if isinstance(value, bool):
        return "ИСТИНА" if value else "ЛОЖЬ"

    text = str(value).strip().casefold()
    if text in {"true", "1", "да", "истина"}:
        return "ИСТИНА"
    if text in {"false", "0", "нет", "ложь"}:
        return "ЛОЖЬ"
    return str(value)


def work_type_number(value):
    """В Excel: 0 — ремонт, 1 — обратное направление."""
    value = clean_value(value)
    if value is None:
        return None
    if isinstance(value, bool):
        return 1 if value else 0
    text = str(value).strip().casefold()
    if text in {"true", "1", "да", "истина"}:
        return 1
    if text in {"false", "0", "нет", "ложь"}:
        return 0
    return value


def display_value(layer, index, foreign_id, *field_candidates, default=None):
    feature = lookup_feature(index, foreign_id)
    if feature is None:
        return default
    return attr(layer, feature, *field_candidates, default=default)


def split_connections(value):
    value = clean_value(value)
    if value is None:
        return []
    parts = re.split(r"\s*;\s*", str(value))
    return [part.strip() for part in parts if part and part.strip()]


_CONNECTION_KM_FIRST = re.compile(
    r"^(?:(?:ТСГ|ГРС|ГИС|БП|ПХГ)\s+)?"
    r"(?P<km>-?\d+(?:[.,]\d+)?)\s*км\s+"
    r"(?P<pipeline>.+?)\s*$",
    flags=re.IGNORECASE,
)

_CONNECTION_PIPELINE_FIRST = re.compile(
    r"^(?P<pipeline>.+?)\s*\+\s*"
    r"(?P<km>-?\d+(?:[.,]\d+)?)\s*км\s*$",
    flags=re.IGNORECASE,
)


def parse_connection(value):
    """
    Возвращает (газопровод, километр, исходный_текст).
    При невозможности разобрать газопровод/километр сохраняет исходный текст.
    """
    if value is None:
        return None, None, None

    text = str(value).strip()
    if not text:
        return None, None, None

    match = _CONNECTION_KM_FIRST.match(text)
    if not match:
        match = _CONNECTION_PIPELINE_FIRST.match(text)

    if not match:
        return None, None, text

    km = as_float(match.group("km"))
    pipeline = match.group("pipeline").strip()
    return pipeline, km, text


def choose_connection(connection_text, target_name=None):
    """
    Выбирает из текста через ';' подключение, наиболее похожее на название ТСГ.
    """
    parts = split_connections(connection_text)
    if not parts:
        return None

    if not target_name:
        return parts[0]

    target = re.sub(r"\s+", " ", str(target_name).strip()).casefold()

    for part in parts:
        normalized = re.sub(r"\s+", " ", part).casefold()
        if normalized == target:
            return part

    for part in parts:
        normalized = re.sub(r"\s+", " ", part).casefold()
        if target in normalized or normalized in target:
            return part

    return parts[0]


def date_difference_days(start_value, end_value):
    """Разница дат в днях либо None."""
    start_value = clean_value(start_value)
    end_value = clean_value(end_value)

    if isinstance(start_value, datetime):
        start_value = start_value.date()
    if isinstance(end_value, datetime):
        end_value = end_value.date()

    if not isinstance(start_value, date) or not isinstance(end_value, date):
        return None

    return (end_value - start_value).days


def period_sort_key(value):
    value = clean_value(value)
    if value is None:
        return (0, "")
    return (1, str(value))


def build_period_routes(period_edges, point_types, add_error):
    """
    Разбивает периодные сегменты на логические маршруты.

    Главный принцип версии 4:
    маршрут строится не до самого дальнего тупика, а между граничными точками.

    Граничная точка — это:
    1) начало или конец сети;
    2) узел ветвления или соединения;
    3) значимый объект: КС, ГРС, ГИС, ПХГ, БП или ТСГ.

    Поэтому длинная цепочка:
        КС-1 -> техническая точка -> КС-2 -> техническая точка -> КС-3
    превращается в два маршрута:
        КС-1 -> КС-2
        КС-2 -> КС-3

    Повторение одинакового названия точки разрешено.
    Уникальность определяется UUID, а не текстом названия.
    """
    groups = defaultdict(list)

    for edge in period_edges:
        key = (
            uuid_key(edge["pipeline_id"]),
            edge["start_date"],
            edge["end_date"],
            edge["work_type"],
            str(edge["remark"] or "").strip(),
        )
        groups[key].append(edge)

    routes = []

    def edge_key(edge):
        period_id = uuid_key(edge["period_id"])
        if period_id:
            return period_id
        return (
            f"{uuid_key(edge['segment_id'])}|"
            f"{edge['start_date']}|{edge['end_date']}|"
            f"{edge['work_type']}|{edge['remark']}"
        )

    def types_text(node_id):
        values = sorted(point_types.get(node_id, set()))
        return ", ".join(values)

    for group_key, edges in groups.items():
        by_start = defaultdict(list)
        by_end = defaultdict(list)
        edge_by_key = {}

        for edge in edges:
            start_key = uuid_key(edge["start_id"])
            end_key = uuid_key(edge["end_id"])
            by_start[start_key].append(edge)
            by_end[end_key].append(edge)
            edge_by_key[edge_key(edge)] = edge

        def edge_sort_key(edge):
            km = as_float(edge.get("start_km"))
            return (
                km if km is not None else 0,
                str(edge.get("start_name") or ""),
                str(edge.get("end_name") or ""),
                str(edge.get("period_id") or ""),
            )

        for values in by_start.values():
            values.sort(key=edge_sort_key)

        nodes = set(by_start) | set(by_end)

        def is_boundary(node_id):
            if node_id is None:
                return True
            in_degree = len(by_end.get(node_id, []))
            out_degree = len(by_start.get(node_id, []))
            return (
                in_degree != 1
                or out_degree != 1
                or bool(point_types.get(node_id))
            )

        def boundary_reason(node_id):
            reasons = []
            in_degree = len(by_end.get(node_id, []))
            out_degree = len(by_start.get(node_id, []))
            node_types = types_text(node_id)

            if node_types:
                reasons.append(f"значимый объект: {node_types}")
            if in_degree == 0:
                reasons.append("начало сети")
            if out_degree == 0:
                reasons.append("конец сети")
            if in_degree > 1 and out_degree > 1:
                reasons.append("узел соединения и ветвления")
            elif in_degree > 1:
                reasons.append("узел соединения")
            elif out_degree > 1:
                reasons.append("узел ветвления")

            return "; ".join(reasons) or "граница маршрута"

        boundary_nodes = {node for node in nodes if is_boundary(node)}
        covered = set()
        group_paths = []

        def save_path(path, start_reason=None, end_reason=None):
            if not path:
                return

            signature = tuple(edge_key(edge) for edge in path)
            if signature in {
                tuple(edge_key(edge) for edge in existing["path"])
                for existing in group_paths
            }:
                return

            for key in signature:
                covered.add(key)

            group_paths.append({
                "path": path,
                "start_reason": start_reason or boundary_reason(
                    uuid_key(path[0]["start_id"])
                ),
                "end_reason": end_reason or boundary_reason(
                    uuid_key(path[-1]["end_id"])
                ),
            })

        # Маршруты, начинающиеся в каждой граничной точке.
        ordered_boundaries = sorted(
            boundary_nodes,
            key=lambda node: (
                str(
                    by_start.get(node, [{}])[0].get("start_name", "")
                    if by_start.get(node)
                    else ""
                ),
                str(node or ""),
            ),
        )

        for start_node in ordered_boundaries:
            for first_edge in by_start.get(start_node, []):
                first_key = edge_key(first_edge)
                if first_key in covered:
                    continue

                path = [first_edge]
                used = {first_key}
                current_node = uuid_key(first_edge["end_id"])

                while current_node not in boundary_nodes:
                    candidates = [
                        candidate
                        for candidate in by_start.get(current_node, [])
                        if edge_key(candidate) not in used
                    ]

                    if len(candidates) != 1:
                        break

                    next_edge = candidates[0]
                    next_key = edge_key(next_edge)
                    path.append(next_edge)
                    used.add(next_key)
                    current_node = uuid_key(next_edge["end_id"])

                save_path(
                    path,
                    start_reason=boundary_reason(start_node),
                    end_reason=boundary_reason(current_node),
                )

        # Остатки: циклы или компоненты без естественной границы.
        for first_edge in sorted(edges, key=edge_sort_key):
            first_key = edge_key(first_edge)
            if first_key in covered:
                continue

            path = []
            used = set()
            current_edge = first_edge

            while current_edge is not None:
                current_key = edge_key(current_edge)
                if current_key in used:
                    break

                used.add(current_key)
                path.append(current_edge)
                next_node = uuid_key(current_edge["end_id"])

                if next_node in boundary_nodes and len(path) > 0:
                    break

                candidates = [
                    candidate
                    for candidate in by_start.get(next_node, [])
                    if edge_key(candidate) not in used
                ]

                current_edge = candidates[0] if len(candidates) == 1 else None

            save_path(
                path,
                start_reason="цикл или компонент без естественного начала",
                end_reason="цикл или компонент без естественного конца",
            )

        # Контроль: каждый сегмент группы должен войти хотя бы в один маршрут.
        not_covered = [
            key for key in edge_by_key
            if key not in covered
        ]
        if not_covered:
            add_error(
                "Периоды",
                group_key[0] or "",
                "Не включены сегменты",
                f"В маршруты не включено сегментов: {len(not_covered)}.",
            )

        for route_number, item in enumerate(group_paths, start=1):
            path = item["path"]
            first_edge = path[0]
            last_edge = path[-1]
            start_key = uuid_key(first_edge["start_id"])
            end_key = uuid_key(last_edge["end_id"])

            routes.append({
                "route_number_in_group": route_number,
                "pipeline_id": first_edge["pipeline_id"],
                "pipeline_name": first_edge["pipeline_name"],
                "start_id": first_edge["start_id"],
                "end_id": last_edge["end_id"],
                "start_name": first_edge["start_name"],
                "end_name": last_edge["end_name"],
                "start_types": types_text(start_key),
                "end_types": types_text(end_key),
                "start_reason": item["start_reason"],
                "end_reason": item["end_reason"],
                "start_date": first_edge["start_date"],
                "end_date": first_edge["end_date"],
                "work_type": first_edge["work_type"],
                "remark": first_edge["remark"],
                "segment_count": len(path),
                "period_ids": [edge["period_id"] for edge in path],
                "segment_ids": [edge["segment_id"] for edge in path],
            })

    routes.sort(
        key=lambda route: (
            period_sort_key(route["start_date"]),
            period_sort_key(route["end_date"]),
            str(route["pipeline_name"] or ""),
            str(route["start_name"] or ""),
            str(route["end_name"] or ""),
            route["route_number_in_group"],
        )
    )
    return routes


def build_work_mode_routes(period_routes):
    """
    Формирует лист 'Режим работы' как перечень уникальных логических маршрутов.

    В отличие от 'ЗимаЛето', здесь один и тот же маршрут не повторяется
    для каждого сезонного периода. Уникальность:
    газопровод + UUID начала + UUID конца + WorkType + Remark.

    Если у маршрута несколько разных пар дат, даты на листе
    'Режим работы' оставляются пустыми. Все периоды остаются видны
    на листе 'ЗимаЛето' и на контрольном листе.
    """
    grouped = {}

    for route in period_routes:
        key = (
            uuid_key(route["pipeline_id"]),
            uuid_key(route["start_id"]),
            uuid_key(route["end_id"]),
            route["work_type"],
            str(route["remark"] or "").strip(),
        )

        if key not in grouped:
            grouped[key] = {
                **route,
                "period_pairs": set(),
                "occurrence_count": 0,
            }

        grouped[key]["period_pairs"].add(
            (route["start_date"], route["end_date"])
        )
        grouped[key]["occurrence_count"] += 1

    result = []

    for route in grouped.values():
        pairs = route["period_pairs"]

        if len(pairs) == 1:
            start_date, end_date = next(iter(pairs))
        else:
            start_date, end_date = None, None

        item = dict(route)
        item["start_date"] = start_date
        item["end_date"] = end_date
        item["period_variant_count"] = len(pairs)
        result.append(item)

    result.sort(
        key=lambda route: (
            str(route["pipeline_name"] or ""),
            str(route["start_name"] or ""),
            str(route["end_name"] or ""),
            str(route["remark"] or ""),
        )
    )
    return result


def create_sheet(workbook, name, title, headers):
    worksheet = workbook.create_sheet(title=name)
    worksheet.sheet_view.showGridLines = False

    last_col = get_column_letter(len(headers))
    worksheet.merge_cells(f"A1:{last_col}1")
    worksheet["A1"] = title
    worksheet["A1"].font = Font(name="Calibri", size=13, bold=True)
    worksheet["A1"].alignment = Alignment(horizontal="center", vertical="center")
    worksheet.row_dimensions[1].height = 24

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    thin = Side(style="thin", color="B7C9D6")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for column, header in enumerate(headers, start=1):
        cell = worksheet.cell(row=2, column=column, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )
        cell.border = border

    worksheet.freeze_panes = "A3"
    worksheet.row_dimensions[2].height = 42
    return worksheet


def append_row(worksheet, values):
    worksheet.append([clean_value(value) for value in values])


def finalize_sheet(worksheet, max_width=42):
    """Базовое форматирование после заполнения."""
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    max_row = worksheet.max_row
    max_column = worksheet.max_column

    if max_row >= 2 and max_column >= 1:
        worksheet.auto_filter.ref = (
            f"A2:{get_column_letter(max_column)}{max_row}"
        )

    # Ограничиваем расчёт ширины первыми 1000 строками.
    scan_to = min(max_row, 1002)
    for column in range(1, max_column + 1):
        width = 10
        for row in range(1, scan_to + 1):
            value = worksheet.cell(row=row, column=column).value
            if value is None:
                continue
            length = max(
                (len(line) for line in str(value).splitlines()),
                default=0,
            )
            width = max(width, min(length + 2, max_width))
        worksheet.column_dimensions[get_column_letter(column)].width = width

    for row in worksheet.iter_rows(min_row=3, max_row=max_row):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border

            if isinstance(cell.value, datetime):
                cell.number_format = "DD.MM.YYYY HH:MM"
            elif isinstance(cell.value, date):
                cell.number_format = "DD.MM.YYYY"

    # Немного сужаем служебный первый столбец.
    worksheet.column_dimensions["A"].width = min(
        worksheet.column_dimensions["A"].width or 14,
        22,
    )


def main():
    project = QgsProject.instance()
    if not project.mapLayers():
        QMessageBox.critical(
            None,
            "Нет слоёв",
            "В текущем проекте QGIS нет загруженных слоёв или таблиц."
        )
        return

    layers = {name: get_layer(name) for name in EXPECTED_LAYERS}
    missing = [name for name, layer in layers.items() if layer is None]
    missing_critical = [name for name in CRITICAL_LAYERS if layers[name] is None]

    if missing_critical:
        QMessageBox.critical(
            None,
            "Не хватает основных таблиц",
            "Не найдены критичные таблицы/слои:\n\n"
            + "\n".join(missing_critical)
            + "\n\nЗагрузите их в текущий проект QGIS и повторите запуск."
        )
        return

    desktop = QStandardPaths.writableLocation(QStandardPaths.DesktopLocation)
    default_path = str(Path(desktop) / "Выгрузка_ГТЛ_из_QGIS_v4.xlsx")

    output_path, _ = QFileDialog.getSaveFileName(
        None,
        "Сохранить выгрузку Excel",
        default_path,
        "Excel (*.xlsx)",
    )
    if not output_path:
        return
    if not output_path.lower().endswith(".xlsx"):
        output_path += ".xlsx"

    errors = []

    def add_error(section, object_id, error_type, message):
        errors.append([
            section,
            object_id,
            error_type,
            message,
        ])

    for name in missing:
        add_error(
            "Структура проекта",
            name,
            "Нет слоя",
            "Таблица/слой не загружен в проект. Соответствующий лист может быть пустым.",
        )

    # Зафиксированная логика третьей версии.
    add_error(
        "Правила выгрузки",
        "",
        "ТСГ",
        "Название месторождения/полевого объекта берётся из GtsPoints.Name, "
        "GUID — из FieldPoints.Id.",
    )
    add_error(
        "Правила выгрузки",
        "",
        "Лицензия",
        "Номер лицензии берётся только из LicUch.lic. "
        "Пустые LuId и lic сохраняются пустыми.",
    )
    add_error(
        "Правила выгрузки",
        "",
        "Периоды",
        "'Режим работы' содержит уникальные логические маршруты без "
        "повтора по сезонным периодам. 'ЗимаЛето' содержит все варианты "
        "маршрутов по датам, включая записи без дат. Маршрут разрывается "
        "на КС, ГРС, ГИС, ПХГ, БП, ТСГ, а также на ветвлениях и соединениях. "
        "ConnectionPoints маршрут не разрывают. UUID важнее названия.",
    )
    add_error(
        "Правила выгрузки",
        "",
        "Сегменты",
        "Для сегментов нулевой длины в колонку перемычек ставится 1.",
    )
    add_error(
        "Правила выгрузки",
        "",
        "Зоны выхода",
        "Наименование зоны выхода берётся из OutZones.Name.",
    )
    add_error(
        "Правила выгрузки",
        "",
        "ПХГ",
        "Поля магистрального газопровода, километра подключения и "
        "расстояния до МГ сохраняются пустыми, так как источника в БД нет.",
    )

    # Индексы.
    indexes = {
        name: build_index(layer)
        for name, layer in layers.items()
        if layer is not None
    }

    gts_layer = layers["GtsPoints"]
    gts_index = indexes.get("GtsPoints", {})
    region_layer = layers["Regions"]
    region_index = indexes.get("Regions", {})
    transgas_layer = layers["TransGases"]
    transgas_index = indexes.get("TransGases", {})
    inzone_layer = layers["InZones"]
    inzone_index = indexes.get("InZones", {})
    outzone_layer = layers["OutZones"]
    outzone_index = indexes.get("OutZones", {})
    pipeline_layer = layers["GasPipelines"]
    pipeline_index = indexes.get("GasPipelines", {})

    fieldpoints_layer = layers["FieldPoints"]
    fieldpoints_index = indexes.get("FieldPoints", {})

    # Типы значимых объектов используются только для разбиения маршрутов.
    # ConnectionPoints намеренно не включены: это технические точки,
    # через которые маршрут должен продолжаться.
    point_types = defaultdict(set)
    significant_layers = {
        "CompressionStations": "КС",
        "GasDistributionStations": "ГРС",
        "GasMeasuringStations": "ГИС",
        "UndergroundGasStorages": "ПХГ",
        "BalancePoints": "БП",
        "GasDeliveryPoints": "ТСГ",
    }

    for layer_name, type_name in significant_layers.items():
        layer = layers.get(layer_name)
        if layer is None:
            continue
        for feature in layer.getFeatures():
            point_id = attr(layer, feature, "Id")
            point_key = uuid_key(point_id)
            if point_key:
                point_types[point_key].add(type_name)

    lic_layer = layers["LicUch"]
    lic_index = indexes.get("LicUch", {})
    user_layer = layers["SubsoilUsers"]
    user_index = indexes.get("SubsoilUsers", {})

    def point_data(point_feature):
        if point_feature is None:
            return {
                "id": None,
                "name": None,
                "latitude": None,
                "longitude": None,
                "in_zone": None,
                "out_zone": None,
                "transgas": None,
                "remark": None,
                "updated": None,
            }

        point_id = attr(gts_layer, point_feature, "Id", "id")
        in_zone_id = attr(gts_layer, point_feature, "InZoneId")
        out_zone_id = attr(gts_layer, point_feature, "OutZoneId")
        transgas_id = attr(gts_layer, point_feature, "TransGasId")

        in_zone = display_value(
            inzone_layer,
            inzone_index,
            in_zone_id,
            "Number",
            "Name",
        )
        out_zone = display_value(
            outzone_layer,
            outzone_index,
            out_zone_id,
            "Name",
        )
        transgas = display_value(
            transgas_layer,
            transgas_index,
            transgas_id,
            "Name",
        )

        return {
            "id": point_id,
            "name": attr(gts_layer, point_feature, "Name"),
            "latitude": attr(gts_layer, point_feature, "Latitude"),
            "longitude": attr(gts_layer, point_feature, "Longitude"),
            "in_zone": in_zone,
            "out_zone": out_zone,
            "transgas": transgas,
            "remark": attr(gts_layer, point_feature, "Remark"),
            "updated": attr(gts_layer, point_feature, "UpdatedDate"),
        }

    def region_name(region_id):
        return display_value(
            region_layer,
            region_index,
            region_id,
            "Name",
        )

    def point_name(point_id):
        point = lookup_feature(gts_index, point_id)
        return attr(gts_layer, point, "Name")

    def pipeline_name(pipeline_id):
        return display_value(
            pipeline_layer,
            pipeline_index,
            pipeline_id,
            "Name",
        )

    def field_point_name_and_guid(field_point_id):
        """
        FieldPoints.Id одновременно является ссылкой на GtsPoints.Id.
        Название берём непосредственно из GtsPoints.Name.
        Пустой FieldPointId является допустимым значением.
        """
        if field_point_id is None:
            return None, None

        field_feature = lookup_feature(fieldpoints_index, field_point_id)
        base_point = lookup_feature(gts_index, field_point_id)

        if field_feature is None:
            add_error(
                "Связи объектов",
                field_point_id,
                "Нет FieldPoints",
                "Идентификатор не найден в FieldPoints.",
            )

        if base_point is None:
            add_error(
                "Связи объектов",
                field_point_id,
                "Нет GtsPoints",
                "FieldPoints.Id не найден в GtsPoints.",
            )

        return attr(gts_layer, base_point, "Name"), clean_value(field_point_id)

    # Длины коллекторов по FieldPointId.
    collector_lengths = defaultdict(list)
    collectors_layer = layers["Collectors"]
    if collectors_layer is not None:
        for feature in collectors_layer.getFeatures():
            field_id = attr(collectors_layer, feature, "FieldPointId")
            key = uuid_key(field_id)
            length = attr(collectors_layer, feature, "Length")
            if key and length is not None:
                collector_lengths[key].append(length)

    workbook = Workbook()
    workbook.remove(workbook.active)

    # 1. Сегменты.
    ws = create_sheet(
        workbook,
        "Сегменты",
        "Таблица 1. Сегменты газопроводов",
        [
            "Атрибутивное поле",
            "Наименование",
            "Начало участка, км",
            "Точка начала",
            "Длина, км",
            "Точка конца",
            "Обратное направление",
            "Мощность, млрд. куб. м в год",
            "Давление, МПа",
            "Диаметр труб, мм",
            "Состояние газопровода",
            "Принадлежность к ТрансГазу",
            "Примечание",
            "Субъект РФ",
            "Точки перемычки, которые не должны игнорироваться (помечены единицей)",
        ],
    )

    segment_layer = layers["GasPipeSegments"]
    segment_features = list(segment_layer.getFeatures())
    segment_features.sort(
        key=lambda f: (
            str(pipeline_name(attr(segment_layer, f, "GasPipelineId")) or ""),
            as_float(attr(segment_layer, f, "StartKilometer")) or 0,
        )
    )

    for feature in segment_features:
        segment_id = attr(segment_layer, feature, "Id")
        start_id = attr(segment_layer, feature, "StartGtsPointId")
        end_id = attr(segment_layer, feature, "EndGtsPointId")
        pipeline_id = attr(segment_layer, feature, "GasPipelineId")
        transgas_id = attr(segment_layer, feature, "TransGasId")
        region_id = attr(segment_layer, feature, "RegionId")
        length = attr(segment_layer, feature, "Length")

        if not point_name(start_id):
            add_error("Сегменты", segment_id, "Нет точки начала", str(start_id))
        if not point_name(end_id):
            add_error("Сегменты", segment_id, "Нет точки конца", str(end_id))

        append_row(
            ws,
            [
                None,
                pipeline_name(pipeline_id),
                attr(segment_layer, feature, "StartKilometer"),
                point_name(start_id),
                length,
                point_name(end_id),
                None,
                attr(segment_layer, feature, "Power"),
                attr(segment_layer, feature, "Pressure"),
                attr(segment_layer, feature, "Diameter"),
                attr(segment_layer, feature, "State"),
                display_value(
                    transgas_layer,
                    transgas_index,
                    transgas_id,
                    "Name",
                ),
                attr(segment_layer, feature, "Remark"),
                region_name(region_id),
                1 if is_zero(length) else None,
            ],
        )
    finalize_sheet(ws)

    # 2. Координаты отдельных объектов.
    ws = create_sheet(
        workbook,
        "Координаты отдельных объектов",
        "Таблица 11. Координаты точек врезки",
        [
            "Атрибутивное поле",
            "Наименование",
            "Широта WGS84",
            "Долгота WGS84",
        ],
    )
    connection_layer = layers["ConnectionPoints"]
    if connection_layer is not None:
        rows = []
        for feature in connection_layer.getFeatures():
            object_id = attr(connection_layer, feature, "Id")
            point = lookup_feature(gts_index, object_id)
            data = point_data(point)
            if point is None:
                add_error(
                    "Координаты отдельных объектов",
                    object_id,
                    "Нет GtsPoints",
                    "ConnectionPoints.Id не найден в GtsPoints.",
                )
            rows.append(
                [None, data["name"], data["latitude"], data["longitude"]]
            )
        rows.sort(key=lambda row: str(row[1] or ""))
        for row in rows:
            append_row(ws, row)
    finalize_sheet(ws)

    # 3. ТСГ.
    ws = create_sheet(
        workbook,
        "ТСГ",
        "Таблица 3. Точки сдачи газа",
        [
            "Атрибутивное поле",
            "Газопровод",
            "Км подключения",
            "Месторождение",
            "Месторождение GUID",
            "Наименование ТСГ",
            "Недропользователь",
            "Недропользователь GUID",
            "Лицензионный участок",
            "Лицензионный участок GUID",
            "Лицензия",
            "Субъект РФ",
            "Зона входа",
            "Широта WGS84",
            "Долгота WGS84",
            "Длина коллектора, км",
            "Примечание",
            "Дата обновления",
            "Расстояние до МГ, км",
        ],
    )

    assignments_layer = layers["FieldPointAssignments"]
    delivery_layer = layers["GasDeliveryPoints"]
    delivery_index = indexes.get("GasDeliveryPoints", {})

    if assignments_layer is not None:
        tsg_rows = []
        for assignment in assignments_layer.getFeatures():
            assignment_id = attr(assignments_layer, assignment, "Id")
            field_id = attr(assignments_layer, assignment, "FieldPointId")
            delivery_id = attr(
                assignments_layer,
                assignment,
                "GasDeliveryPointId",
            )
            lu_id = attr(assignments_layer, assignment, "LuId")
            user_id = attr(assignments_layer, assignment, "SubsoilUserId")

            # Для листа ТСГ нужны только назначения, связанные с GasDeliveryPoint.
            if delivery_id is None:
                continue

            delivery_feature = lookup_feature(delivery_index, delivery_id)
            tsg_point = lookup_feature(gts_index, delivery_id)
            tsg_data = point_data(tsg_point)

            if delivery_feature is None:
                add_error(
                    "ТСГ",
                    assignment_id,
                    "Нет GasDeliveryPoints",
                    f"GasDeliveryPointId={delivery_id}",
                )
            if tsg_point is None:
                add_error(
                    "ТСГ",
                    assignment_id,
                    "Нет GtsPoints для ТСГ",
                    f"GasDeliveryPointId={delivery_id}",
                )

            field_feature = lookup_feature(fieldpoints_index, field_id)
            connection_text = attr(
                fieldpoints_layer,
                field_feature,
                "ConnectionKilometer",
            )
            selected_connection = choose_connection(
                connection_text,
                tsg_data["name"],
            )
            gas_pipeline, connection_km, raw_connection = parse_connection(
                selected_connection
            )

            if raw_connection and gas_pipeline is None:
                add_error(
                    "ТСГ",
                    assignment_id,
                    "Не разобрано подключение",
                    raw_connection,
                )

            field_name, field_guid = field_point_name_and_guid(field_id)

            lic_feature = lookup_feature(lic_index, lu_id)
            license_area = attr(lic_layer, lic_feature, "naim", "Name")
            license_value = attr(lic_layer, lic_feature, "lic")

            if lu_id is not None and lic_feature is None:
                add_error(
                    "ТСГ",
                    assignment_id,
                    "Нет LicUch",
                    f"LuId={lu_id}",
                )
            elif lu_id is not None and license_value in (None, ""):
                add_error(
                    "ТСГ",
                    assignment_id,
                    "Пустой номер лицензии",
                    f"Для лицензионного участка {lu_id} поле LicUch.lic пустое.",
                )

            user_feature = lookup_feature(user_index, user_id)
            user_name = attr(user_layer, user_feature, "Name")

            lengths = collector_lengths.get(uuid_key(field_id), [])
            collector_length_text = (
                "; ".join(str(clean_value(value)) for value in lengths)
                if lengths
                else None
            )

            tsg_rows.append(
                [
                    None,
                    gas_pipeline,
                    connection_km,
                    field_name,
                    field_guid,
                    tsg_data["name"],
                    user_name,
                    clean_value(user_id),
                    license_area,
                    clean_value(lu_id),
                    license_value,
                    region_name(
                        attr(delivery_layer, delivery_feature, "RegionId")
                    ),
                    tsg_data["in_zone"],
                    tsg_data["latitude"],
                    tsg_data["longitude"],
                    collector_length_text,
                    tsg_data["remark"],
                    tsg_data["updated"],
                    None,
                ]
            )

        tsg_rows.sort(
            key=lambda row: (
                str(row[1] or ""),
                as_float(row[2]) or 0,
                str(row[3] or ""),
            )
        )
        for row in tsg_rows:
            append_row(ws, row)
    finalize_sheet(ws)

    # Универсальная функция для специализированной точки.
    def specialized_rows(layer_name):
        layer = layers[layer_name]
        if layer is None:
            return []
        result = []
        for feature in layer.getFeatures():
            object_id = attr(layer, feature, "Id")
            point = lookup_feature(gts_index, object_id)
            if point is None:
                add_error(
                    layer_name,
                    object_id,
                    "Нет GtsPoints",
                    f"{layer_name}.Id не найден в GtsPoints.",
                )
            result.append((feature, point_data(point)))
        return result

    # 4. КС.
    ws = create_sheet(
        workbook,
        "КС",
        "Таблица 2. Компрессорные станции",
        [
            "Атрибутивное поле",
            "Наименование",
            "Субъект РФ",
            "Трансгаз",
            "Широта WGS84",
            "Долгота WGS84",
            "Зона входа",
            "Зона выхода",
            "Мощность, МВт",
            "Примечание",
        ],
    )
    ks_layer = layers["CompressionStations"]
    rows = []
    for feature, data in specialized_rows("CompressionStations"):
        rows.append(
            [
                None,
                data["name"],
                region_name(attr(ks_layer, feature, "RegionId")),
                data["transgas"],
                data["latitude"],
                data["longitude"],
                data["in_zone"],
                data["out_zone"],
                attr(ks_layer, feature, "Power"),
                data["remark"],
            ]
        )
    rows.sort(key=lambda row: str(row[1] or ""))
    for row in rows:
        append_row(ws, row)
    finalize_sheet(ws)

    # 5. ГРС.
    ws = create_sheet(
        workbook,
        "ГРС",
        "Таблица 4. Газораспределительные станции",
        [
            "Атрибутивное поле (ID УМИФГР)",
            "Наименование",
            "Субъект РФ",
            "Населённый пункт",
            "Принадлежность к ТрансГазу",
            "Наименование МГ",
            "Километр подключения к МГ",
            "Расстояние от МГ до ГРС, км",
            "Широта WGS84",
            "Долгота WGS84",
            "Зона выхода",
            "Проектная мощность, тыс. м. куб./час",
            "Текущая загрузка ГРС, тыс. м. куб./час",
            "Независимость (Истина/Ложь)",
            "Примечание",
            "КМ+ГП",
            "кол-во",
        ],
    )
    grs_layer = layers["GasDistributionStations"]
    grs_rows = []
    for feature, data in specialized_rows("GasDistributionStations"):
        connection_text = attr(grs_layer, feature, "ConnectionKilometer")
        parts = split_connections(connection_text) or [None]

        for part in parts:
            gas_pipeline, connection_km, raw_connection = parse_connection(part)
            if raw_connection and gas_pipeline is None:
                add_error(
                    "ГРС",
                    attr(grs_layer, feature, "Id"),
                    "Не разобрано подключение",
                    raw_connection,
                )

            grs_rows.append(
                [
                    attr(grs_layer, feature, "IdUMIFR"),
                    data["name"],
                    region_name(attr(grs_layer, feature, "RegionId")),
                    attr(grs_layer, feature, "Locality"),
                    data["transgas"],
                    gas_pipeline,
                    connection_km,
                    attr(
                        grs_layer,
                        feature,
                        "DistanceFromMainGasPipeline",
                    ),
                    data["latitude"],
                    data["longitude"],
                    data["out_zone"],
                    attr(grs_layer, feature, "DesignCapacity"),
                    attr(grs_layer, feature, "Loading"),
                    bool_ru(attr(grs_layer, feature, "IsIndependent")),
                    data["remark"],
                    raw_connection,
                    1,
                ]
            )

    grs_rows.sort(
        key=lambda row: (
            str(row[5] or ""),
            as_float(row[6]) or 0,
            str(row[1] or ""),
        )
    )
    for row in grs_rows:
        append_row(ws, row)
    finalize_sheet(ws)

    # 6. ГИС.
    ws = create_sheet(
        workbook,
        "ГИС",
        "Таблица 5. Газоизмерительные станции",
        [
            "Атрибутивное поле",
            "Наименование",
            "Субъект РФ",
            "Широта WGS84",
            "Долгота WGS84",
            "Примечание",
            "Зона входа",
            "Зона выхода",
            "Газопровод подключения",
            "км подключения к МГ",
        ],
    )
    gis_layer = layers["GasMeasuringStations"]
    gis_rows = []
    for feature, data in specialized_rows("GasMeasuringStations"):
        connection_text = attr(gis_layer, feature, "ConnectionKilometer")
        parts = split_connections(connection_text) or [None]
        for part in parts:
            gas_pipeline, connection_km, raw_connection = parse_connection(part)
            if raw_connection and gas_pipeline is None:
                add_error(
                    "ГИС",
                    attr(gis_layer, feature, "Id"),
                    "Не разобрано подключение",
                    raw_connection,
                )
            gis_rows.append(
                [
                    None,
                    data["name"],
                    region_name(attr(gis_layer, feature, "RegionId")),
                    data["latitude"],
                    data["longitude"],
                    data["remark"],
                    data["in_zone"],
                    data["out_zone"],
                    gas_pipeline,
                    connection_km,
                ]
            )
    gis_rows.sort(key=lambda row: str(row[1] or ""))
    for row in gis_rows:
        append_row(ws, row)
    finalize_sheet(ws)

    # 7. ПХГ.
    ws = create_sheet(
        workbook,
        "ПХГ",
        "Таблица 6. Подземные хранилища газа",
        [
            "Атрибутивное поле",
            "Наименование",
            "Субъект РФ",
            "Трансгаз",
            "Широта WGS84",
            "Долгота WGS84",
            "Зона входа",
            "Зона выхода",
            "Магистральный газопровод",
            "Км подключения",
            "Расстояние от МГ до ПХГ, км",
            "Точка подключения",
            "Активный объём, млрд. куб. м",
            "Ставка за закачку",
            "Ставка за хранение",
            "Ставка за отбор",
            "Примечание",
        ],
    )
    storage_layer = layers["UndergroundGasStorages"]
    rows = []
    for feature, data in specialized_rows("UndergroundGasStorages"):
        connection_text = attr(
            storage_layer,
            feature,
            "ConnectionKilometer",
        )
        gas_pipeline, connection_km, raw_connection = parse_connection(
            connection_text
        )

        rows.append(
            [
                None,
                data["name"],
                region_name(attr(storage_layer, feature, "RegionId")),
                data["transgas"],
                data["latitude"],
                data["longitude"],
                data["in_zone"],
                data["out_zone"],
                gas_pipeline,
                connection_km,
                None,
                raw_connection,
                attr(storage_layer, feature, "ActiveVolume"),
                attr(storage_layer, feature, "InjectionRate"),
                attr(storage_layer, feature, "AnnualStorageRate"),
                attr(storage_layer, feature, "WithdrawalRate"),
                data["remark"],
            ]
        )
    rows.sort(key=lambda row: str(row[1] or ""))
    for row in rows:
        append_row(ws, row)
    finalize_sheet(ws)

    # 8. БП.
    ws = create_sheet(
        workbook,
        "БП",
        "Таблица 7. Балансовые пункты",
        [
            "Атрибутивное поле",
            "Наименование",
            "Субъект РФ",
            "Трансгаз",
            "Широта WGS84",
            "Долгота WGS84",
            "Зона входа",
            "Зона выхода",
            "Газопровод подключения",
            "км подключения к МГ",
            "Примечание",
        ],
    )
    bp_layer = layers["BalancePoints"]
    bp_rows = []
    for feature, data in specialized_rows("BalancePoints"):
        connection_text = attr(bp_layer, feature, "ConnectionKilometer")
        parts = split_connections(connection_text) or [None]
        for part in parts:
            gas_pipeline, connection_km, raw_connection = parse_connection(part)
            if raw_connection and gas_pipeline is None:
                add_error(
                    "БП",
                    attr(bp_layer, feature, "Id"),
                    "Не разобрано подключение",
                    raw_connection,
                )
            bp_rows.append(
                [
                    None,
                    data["name"],
                    region_name(attr(bp_layer, feature, "RegionId")),
                    data["transgas"],
                    data["latitude"],
                    data["longitude"],
                    data["in_zone"],
                    data["out_zone"],
                    gas_pipeline,
                    connection_km,
                    data["remark"],
                ]
            )
    bp_rows.sort(key=lambda row: str(row[1] or ""))
    for row in bp_rows:
        append_row(ws, row)
    finalize_sheet(ws)

    # 9. Коллектора.
    ws = create_sheet(
        workbook,
        "Коллектора",
        "Таблица 9. Коллектора",
        [
            "Атрибутивное поле",
            "Месторождение",
            "Точка, откуда закачивается газ в КС",
            "Точка КС, у которой будет коллектор",
            "Точка, куда будет откачиваться газ из КС",
            "Длина коллектора, км",
            "Примечание",
        ],
    )
    collector_rows = []
    if collectors_layer is not None:
        for feature in collectors_layer.getFeatures():
            collector_id = attr(collectors_layer, feature, "Id")
            field_id = attr(collectors_layer, feature, "FieldPointId")
            # FieldPointId допускает NULL. Тогда ячейка "Месторождение"
            # остаётся пустой без предупреждения.
            field_name, _ = field_point_name_and_guid(field_id)

            collector_rows.append(
                [
                    None,
                    field_name,
                    point_name(
                        attr(collectors_layer, feature, "SourceGtsPointId")
                    ),
                    point_name(
                        attr(collectors_layer, feature, "TargetGtsPointId")
                    ),
                    point_name(
                        attr(collectors_layer, feature, "GasOutPointId")
                    ),
                    attr(collectors_layer, feature, "Length"),
                    attr(collectors_layer, feature, "Remark"),
                ]
            )
    collector_rows.sort(key=lambda row: str(row[1] or ""))
    for row in collector_rows:
        append_row(ws, row)
    finalize_sheet(ws)

    # 10 и 11. Периоды работы.
    work_headers = [
        "Атрибутивное поле",
        "Начало сегмента",
        "Конец сегмента",
        "Время начала",
        "Время конца",
        "Тип работы",
        "Примечание",
    ]

    ws_work = create_sheet(
        workbook,
        "Режим работы",
        "Таблица 10. Режим работы",
        work_headers,
    )
    ws_season = create_sheet(
        workbook,
        "ЗимаЛето",
        "Таблица 10.1. Режим работы летом и зимой у ПХГ",
        work_headers,
    )

    periods_layer = layers["SegmentWorkPeriods"]
    segment_index = indexes.get("GasPipeSegments", {})
    period_edges = []

    if periods_layer is not None:
        for feature in periods_layer.getFeatures():
            period_id = attr(periods_layer, feature, "Id")
            segment_id = attr(
                periods_layer,
                feature,
                "GasPipeSegmentId",
            )
            segment = lookup_feature(segment_index, segment_id)

            if segment is None:
                add_error(
                    "Периоды",
                    period_id,
                    "Нет сегмента",
                    f"GasPipeSegmentId={segment_id}",
                )
                continue

            start_id = attr(segment_layer, segment, "StartGtsPointId")
            end_id = attr(segment_layer, segment, "EndGtsPointId")
            pipeline_id = attr(segment_layer, segment, "GasPipelineId")
            start_date = attr(
                periods_layer,
                feature,
                "StartDateTime",
            )
            end_date = attr(
                periods_layer,
                feature,
                "EndDateTime",
            )

            if (start_date is None) != (end_date is None):
                add_error(
                    "Периоды",
                    period_id,
                    "Заполнена только одна дата",
                    f"StartDateTime={start_date}; EndDateTime={end_date}",
                )

            work_type = work_type_number(
                attr(periods_layer, feature, "WorkType")
            )
            start_km = attr(segment_layer, segment, "StartKilometer")
            length = attr(segment_layer, segment, "Length")

            # WorkType=1 означает обратное направление движения.
            # Геометрия сегмента в БД не меняется, поэтому для маршрута
            # меняем местами начальную и конечную точки.
            if work_type == 1:
                route_start_id = end_id
                route_end_id = start_id
                start_km_number = as_float(start_km)
                length_number = as_float(length)
                route_start_km = (
                    start_km_number + length_number
                    if start_km_number is not None
                    and length_number is not None
                    else start_km
                )
            else:
                route_start_id = start_id
                route_end_id = end_id
                route_start_km = start_km

            period_edges.append({
                "period_id": period_id,
                "segment_id": segment_id,
                "pipeline_id": pipeline_id,
                "pipeline_name": pipeline_name(pipeline_id),
                "start_id": route_start_id,
                "end_id": route_end_id,
                "start_name": point_name(route_start_id),
                "end_name": point_name(route_end_id),
                "start_km": route_start_km,
                "length": length,
                "start_date": start_date,
                "end_date": end_date,
                "work_type": work_type,
                "remark": attr(periods_layer, feature, "Remark"),
            })

    period_routes = build_period_routes(period_edges, point_types, add_error)
    work_mode_routes = build_work_mode_routes(period_routes)

    work_route_count = 0
    season_route_count = 0
    suspicious_periods = defaultdict(int)

    # Режим работы: один логический маршрут без повторения
    # по каждому сезонному интервалу.
    for route in work_mode_routes:
        append_row(
            ws_work,
            [
                None,
                route["start_name"],
                route["end_name"],
                route["start_date"],
                route["end_date"],
                route["work_type"],
                route["remark"],
            ],
        )
        work_route_count += 1

    # ЗимаЛето: все варианты маршрутов по периодам.
    # Записи без дат также не отбрасываются.
    for route in period_routes:
        append_row(
            ws_season,
            [
                None,
                route["start_name"],
                route["end_name"],
                route["start_date"],
                route["end_date"],
                route["work_type"],
                route["remark"],
            ],
        )
        season_route_count += 1

        duration = date_difference_days(
            route["start_date"],
            route["end_date"],
        )
        if duration is not None and (duration < 0 or duration > 370):
            suspicious_periods[
                (
                    route["start_date"],
                    route["end_date"],
                    duration,
                )
            ] += 1

    for (start_value, end_value, duration), route_count in (
        suspicious_periods.items()
    ):
        add_error(
            "Периоды",
            "",
            "Подозрительная длительность",
            f"Период {start_value} — {end_value} "
            f"({duration} дней) используется в маршрутах: {route_count}. "
            "Дата сохранена без изменения.",
        )

    add_error(
        "Периоды",
        "",
        "Информация",
        f"Записей SegmentWorkPeriods: {len(period_edges)}; "
        f"маршрутов по периодам: {len(period_routes)}; "
        f"уникальных маршрутов для 'Режим работы': {len(work_mode_routes)}; "
        f"строк 'ЗимаЛето': {season_route_count}.",
    )

    finalize_sheet(ws_work)
    finalize_sheet(ws_season)

    # Подробный контроль того, почему маршрут начался и закончился.
    ws_routes = create_sheet(
        workbook,
        "Маршруты_контроль",
        "Контроль формирования маршрутов",
        [
            "Набор",
            "Газопровод",
            "UUID начала",
            "Начало",
            "Тип начала",
            "Причина начала",
            "UUID конца",
            "Конец",
            "Тип конца",
            "Причина окончания",
            "Дата начала",
            "Дата конца",
            "WorkType",
            "Примечание",
            "Количество сегментов",
            "UUID сегментов",
        ],
    )

    for route in period_routes:
        append_row(
            ws_routes,
            [
                "ЗимаЛето / период",
                route["pipeline_name"],
                route["start_id"],
                route["start_name"],
                route["start_types"],
                route["start_reason"],
                route["end_id"],
                route["end_name"],
                route["end_types"],
                route["end_reason"],
                route["start_date"],
                route["end_date"],
                route["work_type"],
                route["remark"],
                route["segment_count"],
                "; ".join(
                    str(value)
                    for value in route["segment_ids"]
                    if value is not None
                ),
            ],
        )

    for route in work_mode_routes:
        append_row(
            ws_routes,
            [
                "Режим работы / уникальный маршрут",
                route["pipeline_name"],
                route["start_id"],
                route["start_name"],
                route["start_types"],
                route["start_reason"],
                route["end_id"],
                route["end_name"],
                route["end_types"],
                route["end_reason"],
                route["start_date"],
                route["end_date"],
                route["work_type"],
                route["remark"],
                route["segment_count"],
                "; ".join(
                    str(value)
                    for value in route["segment_ids"]
                    if value is not None
                ),
            ],
        )

    finalize_sheet(ws_routes, max_width=70)

    # Описание источников, связей и условий прямо в итоговом Excel.
    ws_rules = create_sheet(
        workbook,
        "Правила_и_связи",
        "Источники данных, связи и условия выгрузки",
        [
            "Раздел",
            "Источник",
            "Связь / условие",
            "Что получается в Excel",
        ],
    )

    rules_rows = [
        [
            "Маршруты",
            "SegmentWorkPeriods",
            "GasPipeSegmentId -> GasPipeSegments.Id",
            "Период, WorkType и примечание прикрепляются к сегменту.",
        ],
        [
            "Маршруты",
            "GasPipeSegments",
            "StartGtsPointId / EndGtsPointId -> GtsPoints.Id",
            "Определяются начальная и конечная точки каждого сегмента.",
        ],
        [
            "Маршруты",
            "WorkType = 1",
            "Начало и конец сегмента меняются местами только при построении маршрута.",
            "Формируется обратное направление без изменения базы.",
        ],
        [
            "Маршруты",
            "КС, ГРС, ГИС, ПХГ, БП, ТСГ",
            "UUID точки присутствует в соответствующей специализированной таблице.",
            "В такой точке заканчивается один маршрут и начинается следующий.",
        ],
        [
            "Маршруты",
            "ConnectionPoints",
            "Техническая точка не является границей маршрута.",
            "Соседние сегменты продолжают объединяться.",
        ],
        [
            "Маршруты",
            "Граф сегментов",
            "Маршрут разрывается при числе входов или выходов, отличном от 1.",
            "Ветвления и соединения становятся границами маршрутов.",
        ],
        [
            "Режим работы",
            "Маршруты по периодам",
            "Одинаковые UUID начала/конца, газопровод, WorkType и Remark объединяются независимо от дат.",
            "Один логический маршрут без сезонных повторов.",
        ],
        [
            "ЗимаЛето",
            "Маршруты по периодам",
            "Каждая пара StartDateTime / EndDateTime сохраняется отдельно; NULL/NULL не удаляется.",
            "Все сезонные варианты и запись без дат.",
        ],
        [
            "ТСГ",
            "FieldPointAssignments",
            "FieldPointId -> FieldPoints.Id -> GtsPoints.Id",
            "Месторождение и GUID.",
        ],
        [
            "ТСГ",
            "FieldPointAssignments",
            "GasDeliveryPointId -> GasDeliveryPoints.Id -> GtsPoints.Id",
            "Название ТСГ, координаты, зоны и Трансгаз.",
        ],
        [
            "ТСГ",
            "FieldPointAssignments",
            "LuId -> LicUch.id; номер лицензии только из LicUch.lic.",
            "Участок, GUID участка и номер лицензии.",
        ],
        [
            "КС/ГРС/ГИС/ПХГ/БП",
            "Специализированная таблица",
            "Id одновременно ссылается на GtsPoints.Id.",
            "Общие поля берутся из GtsPoints, специальные — из своей таблицы.",
        ],
        [
            "Зоны",
            "InZones / OutZones",
            "InZoneId / OutZoneId -> Id; отображается Name.",
            "Текстовое наименование зоны.",
        ],
        [
            "ГИС/ПХГ/БП",
            "ConnectionKilometer",
            "Поле в базе пустое.",
            "Газопровод и километр подключения остаются пустыми.",
        ],
    ]

    for row in rules_rows:
        append_row(ws_rules, row)

    finalize_sheet(ws_rules, max_width=80)

    # Лист контроля.
    ws_errors = create_sheet(
        workbook,
        "Ошибки_выгрузки",
        "Контроль выгрузки — версия 4",
        [
            "Раздел",
            "Объект / ID",
            "Тип",
            "Сообщение",
        ],
    )

    if not errors:
        append_row(
            ws_errors,
            ["Контроль", "", "Информация", "Ошибок и предупреждений не зафиксировано."],
        )
    else:
        for error_row in errors:
            append_row(ws_errors, error_row)
    finalize_sheet(ws_errors, max_width=65)

    # Убираем потенциально несовместимые символы из имени файла и сохраняем.
    output_path = os.path.normpath(output_path)

    try:
        workbook.save(output_path)
    except PermissionError:
        QMessageBox.critical(
            None,
            "Не удалось сохранить файл",
            "Файл уже открыт в Excel или нет прав на запись:\n\n"
            + output_path
        )
        return
    except Exception as exc:
        log(f"Ошибка сохранения: {exc}", Qgis.Critical)
        QMessageBox.critical(
            None,
            "Ошибка выгрузки",
            f"Не удалось сохранить Excel:\n\n{exc}"
        )
        raise

    log(f"Выгрузка создана: {output_path}", Qgis.Success)
    QMessageBox.information(
        None,
        "Выгрузка завершена",
        "Excel создан.\n\n"
        f"{output_path}\n\n"
        "Сначала проверьте лист 'Ошибки_выгрузки', затем сравните "
        "рабочие листы с исходным файлом. Периоды работы уже объединены "
        "в маршруты."
    )


main()
